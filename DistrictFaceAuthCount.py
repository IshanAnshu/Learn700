import openpyxl
from collections import Counter

def connectExcelInputFile():
    global loc, wb, sheet
    loc = "DistrictFaceAuthCopy.xlsx"
    wb = openpyxl.load_workbook(loc)
    sheet = wb.active
    print('Input Excel File Connected')

def makeListOfDist():
    global list
    cellValue13 = sheet.cell(row=1, column=3).value
    # print(cellValue13)
    # print(type(cellValue13))
    list = [cellValue13.upper()]
    for i in range(2, sheet.max_row):
        cellValues = sheet.cell(row=i, column=3).value
        # print(i,type(cellValues))
        if type(cellValues) == str:
            list.append(cellValues.upper())
        else:
            break
    # print(list)
    print('List Created')

def writeToOutputFile():
    count = Counter(list)
    # print(count)
    wbOutput = openpyxl.Workbook()
    sheetOutput = wbOutput.active
    r = 1
    c = 1
    for dist, distCount in count.items():
        distCell = sheetOutput.cell(row=r, column=1)
        distCell.value = dist
        distCountCell = sheetOutput.cell(row=r, column=2)
        distCountCell.value = distCount
        r = r + 1

        # print(dist, ":", distCount)
    wbOutput.save("DistrictFaceAuthCountOutput.xlsx")
    print('Output File Generated')

def main():
    connectExcelInputFile()
    makeListOfDist()
    writeToOutputFile()

if __name__=='__main__':
    main()
